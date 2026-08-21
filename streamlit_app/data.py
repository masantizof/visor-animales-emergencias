"""Carga y transformación de la matriz de afectación de animales (UNGRD).

Fuente de datos: Google Sheets ("1.AFECTACIONES Y NECESIDADES") publicado en
Google Drive; se descarga como xlsx en cada actualización manual desde la app.
Complementos: GeoJSON de departamentos y municipios (DIVIPOLA) ubicados en el
directorio raíz del proyecto. No se guarda ninguna copia local de los datos:
el visor siempre refleja la última versión publicada en línea.
"""
from __future__ import annotations

import io
import json
import re
import unicodedata
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd
import requests

SOURCE_URL = (
    "https://docs.google.com/spreadsheets/d/"
    "17jQdrq3HXLuVqL0PrYezoXRcV3sYge6Z/export?format=xlsx"
)

# ---------------------------------------------------------------------------
# Columnas fijas de la hoja fuente ("1.AFECTACIONES Y NECESIDADES")
# Fila 1: encabezados A-P (+ banner Q). Fila 2: subencabezados de necesidades Q-V.
# Datos desde fila 3; fila TOTAL ignorada (los totales se recalculan aquí).
# ---------------------------------------------------------------------------
COL_DEPTO = "A"
COL_MUNI = "B"
COL_FUENTE = "P"
COLS_PROD = {"Des": "C", "Les": "D", "Mue": "E", "Res": "F"}
COLS_COMP = {"Des": "G", "Les": "H", "Mue": "I", "Res": "J"}
COL_ALOJ = "K"
COLS_SILV = {"Des": "L", "Les": "M", "Mue": "N", "Res": "O"}
# Par(es) de columnas por necesidad; si hay número en varias, gana la más a la derecha
NEED_COL_PAIRS = {
    "kg_perro": ["Q"],
    "kg_gato": ["R"],
    "med_unid": ["S", "T"],
    "otros_unid": ["U", "V"],
}
DATA_START_ROW = 3

DEPT_FIX = {"COCÓ": "CHOCÓ"}  # typo conocido en la fuente
MUNI_GEO_ALIASES = {
    "BUGA": "GUADALAJARA DE BUGA",
    "CALIMA (DARIEN)": "CALIMA",
    "SANTIAGO DE CALI": "CALI",
}

STATUS_KEYS = ["Des", "Les", "Mue", "Res"]
CATEGORIES = {
    "prod": "Animales de producción",
    "comp": "Animales de compañía",
    "silv": "Animales silvestres",
}
STATUS_LABELS = {"Des": "Desaparecidos", "Les": "Lesionados", "Mue": "Muertos", "Res": "Rescatados"}
DEPT_SHORT = {
    "VALLE DEL CAUCA": "Valle del Cauca",
    "QUINDÍO": "Quindío",
    "RISARALDA": "Risaralda",
    "CHOCÓ": "Chocó",
}


def normalize(s) -> str:
    """Mayúsculas, sin tildes, espacios colapsados — para cruzar contra el GeoJSON."""
    if s is None:
        return ""
    s = str(s).strip().upper()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s)


def _strip_prefix(s: str) -> str:
    """'11. QUIMBAYA' -> 'QUIMBAYA'."""
    m = re.match(r"^\s*(\d+)\s*\.?\s*(.*)$", s or "")
    return m.group(2).strip() if m else (s or "").strip()


def _num(v) -> int:
    if v is None or v == "":
        return 0
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return 0


def _clean_dept(raw) -> str:
    d = str(raw).strip().upper()
    return DEPT_FIX.get(d, str(raw).strip())


# ---------------------------------------------------------------------------
# Descarga y lectura del Excel
# ---------------------------------------------------------------------------
def download_source_xlsx() -> bytes:
    """Descarga la matriz publicada en Google Drive como xlsx."""
    resp = requests.get(
        SOURCE_URL,
        timeout=30,
        headers={"User-Agent": "Mozilla/5.0 (compatible; UNGRD-animal-visor)"},
    )
    ok = resp.status_code == 200 and resp.content[:2] == b"PK"  # xlsx = zip => magic 'PK'
    if not ok:
        raise RuntimeError(
            f"No se pudo descargar la matriz desde Google Drive (HTTP {resp.status_code}). "
            "Verifica que el enlace sea público e intenta de nuevo."
        )
    return resp.content


def _need_value(cells: list) -> tuple[float | None, str | None, bool]:
    """Interpreta las celdas de una necesidad: número (gana la más a la derecha),
    texto libre como detalle, o marca explícita de 'no se requiere apoyo'."""
    num = None
    texts: list[str] = []
    sin_req = False
    for v in reversed(cells):  # derecha -> izquierda: UNIDADES pisa la columna hermana
        if v is None or str(v).strip() == "":
            continue
        s = str(v).strip()
        if "no se requiere" in s.lower():
            sin_req = True
            continue
        try:
            num = float(v)
        except (TypeError, ValueError):
            texts.insert(0, s.lstrip("\n").strip())
    return num, (" · ".join(dict.fromkeys(texts)) if texts else None), sin_req


def read_source_rows(xlsx_bytes: bytes):
    """Recorre la hoja fuente y devuelve (filas_matriz, filas_necesidad) crudas."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    ws = wb.worksheets[0]

    def cell(col: str, r: int):
        return ws[f"{col}{r}"].value

    rows_out, nec_out = [], []
    blank_streak = 0
    r = DATA_START_ROW
    max_row = ws.max_row or (DATA_START_ROW + 2000)

    while r <= max_row:
        depto_raw = cell(COL_DEPTO, r)
        muni_raw = cell(COL_MUNI, r)

        # fila de totales de la hoja fuente: nos detenemos antes
        if muni_raw and str(muni_raw).strip().upper() == "TOTAL":
            break

        if not depto_raw and not muni_raw:
            blank_streak += 1
            if blank_streak >= 4:
                break
            r += 1
            continue
        blank_streak = 0

        if depto_raw and muni_raw:
            dept = _clean_dept(depto_raw)
            muni = _strip_prefix(str(muni_raw))
            fuente_v = cell(COL_FUENTE, r)
            row = {
                "dept": dept,
                "muni": muni,
                "muni_norm": normalize(muni),
                "fuente": (str(fuente_v).strip() if fuente_v else None),
            }
            for k, col in COLS_PROD.items():
                row[f"prod{k}"] = _num(cell(col, r))
            for k, col in COLS_COMP.items():
                row[f"comp{k}"] = _num(cell(col, r))
            for k, col in COLS_SILV.items():
                row[f"silv{k}"] = _num(cell(col, r))
            row["aloj"] = _num(cell(COL_ALOJ, r))
            rows_out.append(row)

            # Necesidades (Q-V) solo tienen sentido con municipio asociado
            nec = {"dept": dept, "muni": muni, "muni_norm": row["muni_norm"]}
            any_num, any_text, sin_req = False, [], False
            for key, cols in NEED_COL_PAIRS.items():
                num, txt, sr = _need_value([cell(c, r) for c in cols])
                nec[key] = num
                if num is not None:
                    any_num = True
                if txt:
                    any_text.append(txt)
                sin_req = sin_req or sr
            nec["otros_text"] = " · ".join(any_text) if any_text else None
            nec["sin_requerimiento"] = sin_req
            nec["reporta"] = any_num or bool(any_text) or sin_req
            if nec["reporta"]:
                nec_out.append(nec)

        r += 1

    wb.close()
    return rows_out, nec_out


def build_registros(rows: list[dict]) -> pd.DataFrame:
    """Agrupa por (depto, municipio) — un municipio puede tener más de una fila fuente."""
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows)
    df["fuente"] = df["fuente"].fillna("")
    num_cols = [c for c in df.columns if c not in ("dept", "muni", "muni_norm", "fuente")]

    agg = {c: "sum" for c in num_cols}
    g = df.groupby(["dept", "muni", "muni_norm"], as_index=False, sort=False).agg(agg)
    fuentes = (
        df[df["fuente"] != ""]
        .groupby(["dept", "muni", "muni_norm"])["fuente"]
        .apply(lambda s: sorted(set(s)))
        .rename("fuentes")
    )
    g = g.merge(fuentes, on=["dept", "muni", "muni_norm"], how="left")
    g["fuentes"] = g["fuentes"].apply(lambda v: v if isinstance(v, list) else [])

    for k in STATUS_KEYS:
        g[f"total_{k}"] = g[f"prod{k}"] + g[f"comp{k}"] + g[f"silv{k}"]
    g["severidad"] = sum(g[f"total_{k}"] for k in STATUS_KEYS)
    g["orden"] = range(len(g))
    return g


def build_necesidades(rows: list[dict]) -> pd.DataFrame:
    """Consolida necesidades por municipio: suma numéricos, une textos libres."""
    if not rows:
        return pd.DataFrame(
            columns=["dept", "muni", "muni_norm", "kg_perro", "kg_gato",
                     "med_unid", "otros_unid", "otros_text", "sin_requerimiento"]
        )
    df = pd.DataFrame(rows)
    agg = df.groupby(["dept", "muni", "muni_norm"], as_index=False, sort=False).agg(
        kg_perro=("kg_perro", lambda s: s.dropna().sum() if s.notna().any() else None),
        kg_gato=("kg_gato", lambda s: s.dropna().sum() if s.notna().any() else None),
        med_unid=("med_unid", lambda s: s.dropna().sum() if s.notna().any() else None),
        otros_unid=("otros_unid", lambda s: s.dropna().sum() if s.notna().any() else None),
        otros_text=("otros_text", lambda s: " · ".join(dict.fromkeys(t for t in s.dropna())) or None),
        sin_requerimiento=("sin_requerimiento", "any"),
    )
    return agg


# ---------------------------------------------------------------------------
# GeoJSON — departamentos y municipios (DIVIPOLA)
# ---------------------------------------------------------------------------
def load_geo(base_dir: Path, target_depts_norm: set[str]):
    with open(base_dir / "Departamentos.geojson", encoding="utf-8") as f:
        dep_geo = json.load(f)
    with open(base_dir / "Municipios.geojson", encoding="utf-8") as f:
        mun_geo = json.load(f)

    dep_feats = [f for f in dep_geo["features"] if normalize(f["properties"]["DPTO_CNMBR"]) in target_depts_norm]
    dep_name_to_code = {normalize(f["properties"]["DPTO_CNMBR"]): f["properties"]["DPTO_CCDGO"] for f in dep_feats}
    dep_code_to_geoname = {f["properties"]["DPTO_CCDGO"]: f["properties"]["DPTO_CNMBR"] for f in dep_feats}

    target_codes = set(dep_name_to_code.values())
    mun_feats = [f for f in mun_geo["features"] if f["properties"]["DPTO_CCDGO"] in target_codes]

    muni_lookup = {}  # (dept_code, norm_name) -> mpio_code
    for f in mun_feats:
        p = f["properties"]
        muni_lookup[(p["DPTO_CCDGO"], normalize(p["MPIO_CNMBR"]))] = p["MPIO_CCNCT"]

    dep_fc = {"type": "FeatureCollection", "features": dep_feats}
    mun_fc = {"type": "FeatureCollection", "features": mun_feats}
    return {
        "dep_geojson": dep_fc,
        "mun_geojson": mun_fc,
        "dep_name_to_code": dep_name_to_code,
        "dep_code_to_geoname": dep_code_to_geoname,
        "muni_lookup": muni_lookup,
        "mun_features_by_dept": {
            code: [f for f in mun_feats if f["properties"]["DPTO_CCDGO"] == code] for code in target_codes
        },
    }


def attach_geo_codes(registros: pd.DataFrame, geo: dict) -> pd.DataFrame:
    registros = registros.copy()
    registros["dept_norm"] = registros["dept"].apply(normalize)
    registros["dept_code"] = registros["dept_norm"].map(geo["dep_name_to_code"])

    def _muni_code(row):
        norm = row["muni_norm"]
        norm = normalize(MUNI_GEO_ALIASES.get(norm, norm))
        return geo["muni_lookup"].get((row["dept_code"], norm))

    registros["muni_code"] = registros.apply(_muni_code, axis=1)
    return registros


# ---------------------------------------------------------------------------
# Punto de entrada
# ---------------------------------------------------------------------------
def load_all(base_dir: Path):
    """Descarga la matriz en línea, la procesa y cruza contra los GeoJSON."""
    xlsx_bytes = download_source_xlsx()
    fetched_at = datetime.now()

    rows, nec_rows = read_source_rows(xlsx_bytes)
    registros = build_registros(rows)
    necesidades = build_necesidades(nec_rows)

    target_depts_norm = set(registros["dept"].apply(normalize).unique()) if not registros.empty else set()
    geo = load_geo(base_dir, target_depts_norm)
    registros = attach_geo_codes(registros, geo)

    fuentes = sorted({f for fl in registros["fuentes"] for f in fl}) if not registros.empty else []

    return {
        "registros": registros,
        "necesidades": necesidades,
        "geo": geo,
        "fuentes": fuentes,
        "fetched_at": fetched_at,
        "n_depts": int(registros["dept"].nunique()) if not registros.empty else 0,
        "n_munis": int(len(registros)) if not registros.empty else 0,
        "source_url": SOURCE_URL,
    }
